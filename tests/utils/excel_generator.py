"""
Excel test-case report generator.

Generates a .xlsx file with one sheet per user story.
Each sheet has a formatted table of test cases.

Usage:
    from utils.excel_generator import generate_test_case_report
    generate_test_case_report(test_cases, output_path)
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_COL_HEADER_BG   = "001B35"   # primary (dark navy)
_COL_HEADER_FG   = "FFFFFF"
_COL_POSITIVE_BG = "E8F5E9"   # light green
_COL_NEGATIVE_BG = "FFEBEE"   # light red
_COL_P0_FG       = "B71C1C"   # dark red for P0
_COL_P1_FG       = "E65100"   # orange for P1
_COL_P2_FG       = "1565C0"   # blue for P2
_COL_BORDER      = "C3C6CF"   # outline-variant

# Column definitions
_COLUMNS = [
    ("Test Case ID",    18),
    ("Title",           38),
    ("AC Reference",    14),
    ("Preconditions",   36),
    ("Test Steps",      52),
    ("Test Data",       28),
    ("Expected Result", 48),
    ("Priority",        10),
    ("Type",            14),
]

_PRIORITY_COLOURS = {
    "P0": _COL_P0_FG,
    "P1": _COL_P1_FG,
    "P2": _COL_P2_FG,
}


def _thin_border() -> Border:
    side = Side(style="thin", color=_COL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font() -> Font:
    return Font(name="Calibri", bold=True, color=_COL_HEADER_FG, size=11)


def _cell_font(bold: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, size=10)


def _wrap_align() -> Alignment:
    return Alignment(wrap_text=True, vertical="top")


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_COL_HEADER_BG)


def _row_fill(test_type: str) -> PatternFill:
    if test_type.lower() == "positive":
        return PatternFill("solid", fgColor=_COL_POSITIVE_BG)
    return PatternFill("solid", fgColor=_COL_NEGATIVE_BG)


def _write_sheet(ws, story_title: str, test_cases: list[dict[str, Any]]) -> None:
    """Populate a single worksheet with test cases."""
    # -- Story title banner -------------------------------------------------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=story_title)
    title_cell.font = Font(name="Calibri", bold=True, size=13, color=_COL_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor="003057")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # -- Header row ---------------------------------------------------------
    for col_idx, (col_name, col_width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[2].height = 22

    # -- Data rows ----------------------------------------------------------
    for row_offset, tc in enumerate(test_cases, start=3):
        row_fill = _row_fill(tc.get("type", ""))
        row_values = [
            tc.get("id", ""),
            tc.get("title", ""),
            tc.get("ac_ref", ""),
            tc.get("preconditions", ""),
            tc.get("steps", ""),
            tc.get("test_data", ""),
            tc.get("expected", ""),
            tc.get("priority", ""),
            tc.get("type", ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = _thin_border()
            cell.alignment = _wrap_align()

            # Priority gets its colour
            if col_idx == 8:  # Priority column
                cell.font = Font(
                    name="Calibri",
                    bold=True,
                    size=10,
                    color=_PRIORITY_COLOURS.get(value, "000000"),
                )
            else:
                cell.font = _cell_font(bold=(col_idx == 1))

        # Auto row height hint (openpyxl can't truly auto-fit; approximate)
        ws.row_dimensions[row_offset].height = 72

    # Freeze panes below header
    ws.freeze_panes = "A3"


def generate_test_case_report(
    stories: dict[str, dict[str, Any]],
    output_path: str | Path,
) -> Path:
    """
    Generate an xlsx report with one sheet per user story.

    Args:
        stories: {
            "SCRUM-2": {
                "title": "SCRUM-2 — Login Screen Implementation",
                "test_cases": [ {...}, ... ]
            },
            ...
        }
        output_path: Destination .xlsx path.

    Returns:
        Path to the created file.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    for story_key, story_data in stories.items():
        sheet_name = story_key[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, story_data["title"], story_data["test_cases"])

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
